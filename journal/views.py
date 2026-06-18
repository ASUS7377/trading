from urllib import request
from django.shortcuts import get_object_or_404
from django.shortcuts import render
from django.shortcuts import redirect
import random
import calendar
from django.http import HttpResponse
from openpyxl import Workbook
from datetime import date
from .models import Profile, TodoTask
from .models import Trade
from django.db.models import Sum
from .forms import ProfileForm
from .forms import TradeForm
from django.contrib.auth import authenticate
from django.contrib.auth import login
from django.contrib.auth import logout
from .forms import LoginForm
from django.shortcuts import render
from django.db.models import Sum
from .models import Trade, Profile
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.shortcuts import get_object_or_404
from django.db.models import Q
from django.core.paginator import Paginator
from django.utils import timezone


@login_required
def dashboard(request):

    profile = Profile.objects.filter(
        user=request.user
    ).first()

    if not profile:

        return redirect(
            'welcome'
        )

    trades = Trade.objects.filter(
        user=request.user
    )

    total_trades = trades.count()

    total_profit = (
        trades.aggregate(
            total=Sum('profit')
        )['total']
        or 0
    )

    winning_trades = trades.filter(
        profit__gt=0
    ).count()

    current_balance = (
        profile.initial_balance
        + total_profit
    )

    win_rate = 0

    if total_trades > 0:

        win_rate = round(
            (
                winning_trades
                / total_trades
            ) * 100,
            1
        )

    context = {

        'profile': profile,

        'current_balance': current_balance,

        'total_profit': total_profit,

        'total_trades': total_trades,

        'win_rate': win_rate,

        'last_trade': trades.first()

    }

    return render(
        request,
        'dashboard.html',
        context
    )

@login_required
def profile_view(request):

    profile = Profile.objects.filter(
        user=request.user
    ).first()

    if not profile:

        return redirect(
            'welcome'
        )

    trades = Trade.objects.filter(
        user=request.user
    )

    total_profit = (
        trades.aggregate(
            total=Sum('profit')
        )['total']
        or 0
    )

    total_trades = trades.count()

    winning_trades = trades.filter(
        profit__gt=0
    ).count()

    win_rate = 0

    if total_trades > 0:

        win_rate = round(
            (
                winning_trades /
                total_trades
            ) * 100,
            2
        )

    current_balance = (
        profile.initial_balance +
        total_profit
    )

    if current_balance >= profile.target_balance:

        goal_completed = True

        remaining_goal = (
            current_balance -
            profile.target_balance
        )

    else:

        goal_completed = False

        remaining_goal = (
            profile.target_balance -
            current_balance
        )

    goal_percent = 0

    if profile.target_balance > 0:

        goal_percent = round(
            (
                current_balance /
                profile.target_balance
            ) * 100,
            1
        )

    if goal_percent > 100:

        goal_percent = 100

    # Уровни трейдера

    if total_profit >= 100000:

        trader_level = 5

        trader_icon = 'bi-crown-fill'

        trader_title = 'Легенда рынка'

        trader_color = '#eab308'

    elif total_profit >= 50000:

        trader_level = 4

        trader_icon = 'bi-gem'

        trader_title = 'Эксперт'

        trader_color = '#7c3aed'

    elif total_profit >= 20000:

        trader_level = 3

        trader_icon = 'bi-fire'

        trader_title = 'Профессиональный трейдер'

        trader_color = '#f97316'

    elif total_profit >= 5000:

        trader_level = 2

        trader_icon = 'bi-lightning-charge-fill'

        trader_title = 'Развивающийся трейдер'

        trader_color = '#2563eb'

    else:

        trader_level = 1

        trader_icon = 'bi-star-fill'

        trader_title = 'Новичок'

        trader_color = '#64748b'

    favorite_symbol = '-'

    symbol_stats = {}

    for trade in trades:

        if trade.symbol not in symbol_stats:

            symbol_stats[
                trade.symbol
            ] = 0

        symbol_stats[
            trade.symbol
        ] += float(
            trade.profit
        )

    if symbol_stats:

        favorite_symbol = max(
            symbol_stats,
            key=symbol_stats.get
        )

    context = {

        'profile':
        profile,

        'current_balance':
        current_balance,

        'remaining_goal':
        remaining_goal,

        'goal_percent':
        goal_percent,

        'goal_completed':
        goal_completed,

        'total_profit':
        total_profit,

        'total_trades':
        total_trades,

        'win_rate':
        win_rate,

        'favorite_symbol':
        favorite_symbol,

        'trader_level':
        trader_level,

        'trader_icon':
        trader_icon,

        'trader_title':
        trader_title,

        'trader_color':
        trader_color,

    }

    return render(
        request,
        'profile.html',
        context
    )

def add_trade(request):

    if request.method == 'POST':

        form = TradeForm(
            request.POST,
            request.FILES
        )

        if form.is_valid():

            trade = form.save(
                commit=False
            )

            trade.trade_type = request.POST.get(
                'trade_type'
            )
            trade.user = request.user

            trade.save()

            return redirect(
                'history'
            )

        else:

            print(
                form.errors
            )

    else:

        form = TradeForm()

    return render(
        request,
        'trade_form.html',
        {
            'form': form
        }
    )






@login_required
def trades_history(request):

    search = request.GET.get(
        'search',
        ''
    )

    trade_type = request.GET.get(
        'type',
        ''
    )

    result = request.GET.get(
        'result',
        ''
    )

    trades = Trade.objects.filter(
        user=request.user
    )

    if search:

        trades = trades.filter(

            symbol__icontains=search

        )

    if trade_type:

        trades = trades.filter(
            trade_type=trade_type
        )

    if result == 'profit':

        trades = trades.filter(
            profit__gt=0
        )

    elif result == 'loss':

        trades = trades.filter(
            profit__lt=0
        )

    trades = trades.order_by(
        '-date',
        '-created_at'
    )

    paginator = Paginator(
        trades,
        8
    )

    page_number = request.GET.get(
        'page'
    )

    trades = paginator.get_page(
        page_number
    )

    return render(
        request,
        'trades_history.html',
        {

            'trades': trades,

            'search': search,

            'trade_type': trade_type,

            'result': result,

        }
    )
@login_required
def calendar_view(request):

    profile = Profile.objects.filter(
        user=request.user
    ).first()

    trades = Trade.objects.filter(
        user=request.user
    ).order_by(
        'date'
    )

    year = request.GET.get(
        'year'
    )

    month = request.GET.get(
        'month'
    )

    if not year:

        year = date.today().year

    else:

        year = int(year)

    if not month:

        month = date.today().month

    else:

        month = int(month)

    prev_month = month - 1
    prev_year = year

    next_month = month + 1
    next_year = year

    if prev_month < 1:

        prev_month = 12
        prev_year -= 1

    if next_month > 12:

        next_month = 1
        next_year += 1

    cal = calendar.monthcalendar(
        year,
        month
    )

    calendar_days = []

    initial_balance = 0

    if profile:

        initial_balance = float(
            profile.initial_balance
        )

    for week in cal:

        week_days = []

        for day in week:

            if day == 0:

                week_days.append(
                    None
                )

                continue

            current_date = date(
                year,
                month,
                day
            )

            day_trades = Trade.objects.filter(
                user=request.user,
                date=current_date
            )

            day_profit = sum(
                float(trade.profit)
                for trade in day_trades
            )

            previous_trades = Trade.objects.filter(
                user=request.user,
                date__lt=current_date
            )

            balance_before_day = (
                initial_balance
                +
                sum(
                    float(trade.profit)
                    for trade in previous_trades
                )
            )

            percent = 0

            if balance_before_day > 0:

                percent = round(
                    (
                        day_profit
                        /
                        balance_before_day
                    ) * 100,
                    2
                )

            week_days.append({

                'day': day,

                'profit': day_profit,

                'percent': percent,

                'today':
                (
                    day == date.today().day
                    and
                    month == date.today().month
                    and
                    year == date.today().year
                )

            })

        calendar_days.append(
            week_days
        )

    total_profit = sum(
        float(t.profit)
        for t in trades
    )

    total_trades = trades.count()

    winning_trades = trades.filter(
        profit__gt=0
    ).count()

    win_rate = 0

    if total_trades > 0:

        win_rate = round(
            (
                winning_trades
                /
                total_trades
            ) * 100,
            1
        )

    affirmations = [

        "Дисциплина важнее прибыли.",

        "Следуй системе, а не эмоциям.",

        "Каждая сделка делает тебя лучше.",

        "Контролируй риск, а не рынок.",

        "Терпение приносит результат.",

        "Каждая сделка — опыт.",

    ]

    context = {

        'profile': profile,

        'calendar_days': calendar_days,

        'year': year,

        'month': month,

        'prev_month': prev_month,

        'prev_year': prev_year,

        'next_month': next_month,

        'next_year': next_year,

        'trades': trades,

        'total_profit': total_profit,

        'total_trades': total_trades,

        'win_rate': win_rate,

        'affirmation': random.choice(
            affirmations
        )

    }

    return render(
        request,
        'calendar.html',
        context
    )
@login_required
def statistics_view(request):

    profile = Profile.objects.filter(
        user=request.user
    ).first()

    if not profile:

        return redirect(
            'welcome'
        )

    trades = Trade.objects.filter(
        user=request.user
    )

    total_profit = sum(
        trade.profit
        for trade in trades
    )

    total_trades = trades.count()

    winning_trades = trades.filter(
        profit__gt=0
    ).count()

    losing_trades = trades.filter(
        profit__lt=0
    ).count()

    buy_count = trades.filter(
        trade_type='BUY'
    ).count()

    sell_count = trades.filter(
        trade_type='SELL'
    ).count()

    tp_count = trades.filter(
        trade_result='TP'
    ).count()

    sl_count = trades.filter(
        trade_result='SL'
    ).count()

    manual_count = trades.filter(
        trade_result='MANUAL'
    ).count()

    current_balance = (
        profile.initial_balance
        + total_profit
    )

    win_rate = 0

    if total_trades > 0:

        win_rate = round(
            (
                winning_trades
                / total_trades
            ) * 100,
            2
        )

    tp_rate = 0

    if total_trades > 0:

        tp_rate = round(
            (
                tp_count
                / total_trades
            ) * 100,
            2
        )

    best_trade = trades.order_by(
        '-profit'
    ).first()

    worst_trade = trades.order_by(
        'profit'
    ).first()

    if total_trades <= 1:

        worst_trade = None

    asset_labels = []

    asset_counts = []

    symbols = trades.values_list(
        'symbol',
        flat=True
    )

    unique_symbols = set(
        symbols
    )

    for symbol in unique_symbols:

        asset_labels.append(
            symbol
        )

        asset_counts.append(

            trades.filter(
                symbol=symbol
            ).count()

        )

    context = {

        'profile': profile,

        'current_balance': current_balance,

        'total_profit': total_profit,

        'total_trades': total_trades,

        'winning_trades': winning_trades,

        'losing_trades': losing_trades,

        'buy_count': buy_count,

        'sell_count': sell_count,

        'tp_count': tp_count,

        'sl_count': sl_count,

        'manual_count': manual_count,

        'tp_rate': tp_rate,

        'win_rate': win_rate,

        'best_trade': best_trade,

        'worst_trade': worst_trade,

        'asset_labels': asset_labels,

        'asset_counts': asset_counts,

    }

    return render(
        request,
        'statistics.html',
        context
    )
def login_view(request):

    if request.user.is_authenticated:

        if request.user.is_superuser:

            return redirect(
                'administrator'
            )

        try:

            profile = Profile.objects.get(
                user=request.user
            )

            if not profile.is_setup_completed:

                return redirect(
                    'welcome'
                )

        except:

            pass

        return redirect(
            'dashboard'
        )

    form = LoginForm(
        request,
        data=request.POST or None
    )

    if request.method == 'POST':

        if form.is_valid():

            user = form.get_user()

            login(
                request,
                user
            )

            try:

                profile = Profile.objects.get(
                    user=user
                )

                if not profile.is_setup_completed:

                    return redirect(
                        'welcome'
                    )

            except:

                pass

            if user.is_superuser:

                return redirect(
                    'administrator'
                )

            return redirect(
                'dashboard'
            )

    return render(
        request,
        'login.html',
        {
            'form': form
        }
    )
def logout_view(request):

    logout(request)

    return redirect(
        'login'
    )


@login_required
def administrator_view(request):

    if not request.user.is_superuser:

        return redirect(
            'dashboard'
        )

    users = User.objects.all()

    total_users = users.count()

    active_users = users.filter(
        is_active=True
    ).count()

    blocked_users = users.filter(
        is_active=False
    ).count()

    context = {

        'users': users,

        'total_users': total_users,

        'active_users': active_users,

        'blocked_users': blocked_users,

    }

    return render(
        request,
        'administrator.html',
        context
    )




@login_required
def create_user_view(request):

    if not request.user.is_superuser:

        return redirect(
            'dashboard'
        )

    if request.method == 'POST':

        username = request.POST.get(
            'username'
        )

        password = request.POST.get(
            'password'
        )

        first_name = request.POST.get(
            'first_name'
        )

        last_name = request.POST.get(
            'last_name'
        )

        user = User.objects.create_user(

            username=username,

            password=password,

            first_name=first_name,

            last_name=last_name

        )

        Profile.objects.create(

            user=user,

            full_name=
            f"{first_name} {last_name}",

            initial_balance=0,

            target_balance=0,

            is_setup_completed=False

        )

        return redirect(
            'administrator'
        )

    return render(
        request,
        'create_user.html'
    )


@login_required
def welcome_view(request):

    profile = Profile.objects.filter(
        user=request.user
    ).first()

    if not profile:

        return redirect(
            'logout'
        )

    if profile.is_setup_completed:

        return redirect(
            'dashboard'
        )

    if request.method == 'POST':

        profile.initial_balance = request.POST.get(
            'initial_balance'
        )

        profile.target_balance = request.POST.get(
            'target_balance'
        )

        if request.FILES.get(
            'avatar'
        ):

            profile.avatar = request.FILES.get(
                'avatar'
            )

        profile.is_setup_completed = True

        profile.save()

        return redirect(
            'dashboard'
        )

    return render(
        request,
        'welcome.html',
        {
            'profile': profile
        }
    )


def trade_detail(request, pk):

    trade = get_object_or_404(
        Trade,
        pk=pk
    )

    return render(
        request,
        'trade_detail.html',
        {
            'trade': trade
        }
    )






@login_required
def edit_trade(request, pk):

    trade = get_object_or_404(
        Trade,
        pk=pk,
        user=request.user
    )

    if request.method == 'POST':

        form = TradeForm(
            request.POST,
            request.FILES,
            instance=trade
        )

        if form.is_valid():

            trade = form.save(
                commit=False
            )

            trade.trade_type = request.POST.get(
                'trade_type'
            )

            trade.user = request.user

            trade.save()

            return redirect(
                'trade_detail',
                pk=trade.pk
            )

    else:

        form = TradeForm(
            instance=trade
        )

    return render(
        request,
        'edit_trade.html',
        {
            'form': form,
            'trade': trade
        }
    )



@login_required
def delete_trade(request, pk):

    trade = get_object_or_404(
        Trade,
        pk=pk,
        user=request.user
    )

    if request.method == 'POST':

        trade.delete()

        return redirect(
            'history'
        )

    return render(
        request,
        'delete_trade.html',
        {
            'trade': trade
        }
    )





@login_required
def user_detail_view(request, user_id):

    if not request.user.is_superuser:

        return redirect(
            'dashboard'
        )

    user = get_object_or_404(
        User,
        id=user_id
    )

    profile = Profile.objects.filter(
        user=user
    ).first()

    trades = Trade.objects.filter(
        user=user
    )

    total_trades = trades.count()

    total_profit = sum(
        trade.profit
        for trade in trades
    )

    context = {

        'selected_user': user,

        'profile': profile,

        'total_trades': total_trades,

        'total_profit': total_profit,

    }

    return render(
        request,
        'user_detail.html',
        context
    )


@login_required
def edit_user_view(request, user_id):

    if not request.user.is_superuser:

        return redirect(
            'dashboard'
        )

    user = get_object_or_404(
        User,
        id=user_id
    )

    if request.method == 'POST':

        user.username = request.POST.get(
            'username'
        )

        user.first_name = request.POST.get(
            'first_name'
        )

        user.last_name = request.POST.get(
            'last_name'
        )

        user.is_active = (
            request.POST.get(
                'is_active'
            ) == 'on'
        )

        password = request.POST.get(
            'password'
        )

        if password:

            user.set_password(
                password
            )

        user.save()

        return redirect(
            'user_detail',
            user_id=user.id
        )

    return render(
        request,
        'edit_user.html',
        {
            'selected_user': user
        }
    )


@login_required
def delete_user_view(request, user_id):

    if not request.user.is_superuser:

        return redirect(
            'dashboard'
        )

    user = get_object_or_404(
        User,
        id=user_id
    )
    if user.is_superuser:

        return redirect(
            'user_detail',
            user_id=user.id
        )
    if user.is_superuser:

        return redirect(
            'administrator'
        )

    if request.method == 'POST':

        user.delete()

        return redirect(
            'administrator'
        )

    return render(
        request,
        'delete_user.html',
        {
            'selected_user': user
        }
    )


@login_required
def update_goal_view(request):

    if request.method == 'POST':

        profile = Profile.objects.get(
            user=request.user
        )

        profile.target_balance = float(
            request.POST.get(
                'target_balance'
            )
        )

        profile.save()

    return redirect(
        'profile'
    )


@login_required
def export_trades_excel(request):

    trades = Trade.objects.filter(
        user=request.user
    )

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = 'Trading Journal'

    headers = [

        'Дата',
        'Актив',
        'Тип сделки',
        'Лот',
        'Цена входа',
        'Цена выхода',
        'Take Profit',
        'Stop Loss',
        'Тип закрытия',
        'Прибыль / Убыток',
        'Комментарий'

    ]

    worksheet.append(headers)

    from openpyxl.styles import Font, PatternFill

    header_fill = PatternFill(
        fill_type="solid",
        start_color="1E293B"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    for cell in worksheet[1]:

        cell.fill = header_fill
        cell.font = header_font

    for trade in trades:

        worksheet.append([

            str(trade.date),

            trade.symbol,

            trade.trade_type,

            trade.lot_size,

            trade.entry_price,

            trade.exit_price,

            trade.take_profit,

            trade.stop_loss,

            trade.get_trade_result_display(),

            trade.profit,

            trade.comment,

        ])

    green_fill = PatternFill(
        fill_type="solid",
        start_color="DCFCE7"
    )

    red_fill = PatternFill(
        fill_type="solid",
        start_color="FEE2E2"
    )

    for row in worksheet.iter_rows(
        min_row=2
    ):

        profit_cell = row[9]

        try:

            if float(profit_cell.value) > 0:

                profit_cell.fill = green_fill

            elif float(profit_cell.value) < 0:

                profit_cell.fill = red_fill

        except:

            pass

    for column in worksheet.columns:

        max_length = 0

        column_letter = column[0].column_letter

        for cell in column:

            try:

                if len(str(cell.value)) > max_length:

                    max_length = len(
                        str(cell.value)
                    )

            except:

                pass

        worksheet.column_dimensions[
            column_letter
        ].width = max_length + 5

    response = HttpResponse(

        content_type=
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    )

    response[
        'Content-Disposition'
    ] = (
        'attachment; filename=Trading_Journal.xlsx'
    )

    workbook.save(
        response
    )

    return response


@login_required
def todo_view(request):

    tasks = TodoTask.objects.filter(
        user=request.user
    ).order_by(
        '-created_at'
    )

    completed_tasks = tasks.filter(
        completed=True
    ).count()

    total_tasks = tasks.count()

    progress = 0

    if total_tasks > 0:

        progress = round(
            (
                completed_tasks /
                total_tasks
            ) * 100
        )

    context = {

        'tasks': tasks,

        'completed_tasks':
        completed_tasks,

        'total_tasks':
        total_tasks,

        'progress':
        progress,

    }

    return render(
        request,
        'todo.html',
        context
    )



@login_required
def add_task(request):

    if request.method == 'POST':

        TodoTask.objects.create(

            user=request.user,

            title=request.POST.get(
                'title'
            ),

            category=request.POST.get(
                'category'
            ),

            deadline=request.POST.get(
                'deadline'
            ) or None

        )

    return redirect(
        'todo'
    )



@login_required
def toggle_task(request, task_id):

    task = get_object_or_404(

        TodoTask,

        id=task_id,

        user=request.user

    )

    task.completed = not task.completed

    if task.completed:

        task.completed_at = timezone.now()

    else:

        task.completed_at = None

    task.save()

    return redirect(
        'todo'
    )
@login_required
def delete_task(request, task_id):

    task = get_object_or_404(

        TodoTask,

        id=task_id,

        user=request.user

    )

    task.delete()

    return redirect(
        'todo'
    )
def landing_page(request):

    if request.user.is_authenticated:

        return redirect(
            'dashboard'
        )

    return render(
        request,
        'landing.html'
    )